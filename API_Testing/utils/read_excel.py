"""
Read excel file data, read in
"""

import API_Testing.utils.project_path
from API_Testing.utils.read_config import ReadConfig
from openpyxl import load_workbook

class Read_the_excel(object):
    # get the excel data
    @staticmethod
    def get_data(data_path, config_path):
        wb = load_workbook(data_path)
        models = eval(ReadConfig.get_config(config_path, 'MODE', 'mode'))

        test_data = []
        for model in models:
            sheet = wb[mode]
            if models[mode] == 'all':
                for row in range(2, sheet.max_row+1):
                    tmp_data = dict()
                    tmp_data['sheet_name'] = model
                    for column in range(1, sheet.max_column+1):
                        tmp_data[sheet.cell(1, column).value] = sheet.cell(row, column).value
                    test_data.append(tmp_data)

            else:
                for case in models[mode]:
                    tmp_data = dict()
                    tmp_data['sheet_name'] = mode
                    for column in range(1, sheet.max_column+1):
                        tmp_data[sheet.cell(1, column).value] = sheet.cell(case_id+1, column).value
                    test_data.append(tmp_data)
        return test_data


